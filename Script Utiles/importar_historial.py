"""
importar_historial.py
══════════════════════════════════════════════════════════════
Importa pagos históricos desde Excel al sistema Alianza.
Los pagos se guardan vinculados al propietario del apartamento
como historial — sin crear ni vincular cargos.

Compatible con la plantilla: plantilla_historial_pagos.xlsx

Uso:
  cd C:\alianza_backend
  python importar_historial.py

O especificando el archivo:
  python importar_historial.py mi_archivo.xlsx
══════════════════════════════════════════════════════════════
"""

import sys, os, time, json
from collections import defaultdict

try:
    import openpyxl
except ImportError:
    print("Falta openpyxl.  Ejecuta: pip install openpyxl"); sys.exit(1)
try:
    import requests
except ImportError:
    print("Falta requests.  Ejecuta: pip install requests"); sys.exit(1)

# ── Configuración ──────────────────────────────────────────
BASE_URL = "https://alianzaresidencial-production.up.railway.app/api"
USERNAME  = "Kadin02"
PASSWORD  = "Panama123*"

DEFAULT_FILE = "plantilla_historial_pagos.xlsx"

# Columnas aceptadas en el Excel (nombres alternativos por si el usuario renombra)
COL_FACTURA    = ["# Factura", "Factura", "N Factura", "factura", "num_factura"]
COL_FECHA      = ["Fecha", "fecha", "Fecha Pago", "fecha_pago"]
COL_APTO       = ["Apartamento", "Apto", "apto", "apartamento", "Unidad"]
COL_RAZON      = ["Razon", "Razón", "razon", "Concepto", "Descripcion"]
COL_MONTO      = ["Cantidad", "Monto", "monto", "cantidad", "Total", "Importe"]
COL_COMENTARIO = ["Comentario", "comentario", "Notas", "Nota", "Observacion"]
COL_PROPIETARIO= ["Propietario", "propietario", "Nombre", "Cliente"]  # opcional


def find_col(headers, candidates):
    """Devuelve el índice de la primera columna que coincida, o None."""
    for i, h in enumerate(headers):
        if str(h or "").strip() in candidates:
            return i
    return None


def safe_float(val):
    if val is None: return None
    try: return round(float(str(val).replace(",", ".")), 2)
    except: return None


def safe_str(val):
    s = str(val or "").strip()
    return s if s not in ("None", "nan", "") else None


def safe_date(val):
    if val is None: return None
    if hasattr(val, "strftime"): return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    # Formatos comunes
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y"):
        try:
            from datetime import datetime
            return datetime.strptime(s[:10], fmt).strftime("%Y-%m-%d")
        except: pass
    return s[:10] if len(s) >= 10 else None


# ── Cliente API ────────────────────────────────────────────
class API:
    def __init__(self):
        self.s = requests.Session()

    def login(self):
        try:
            r = self.s.post(f"{BASE_URL}/auth/login",
                            json={"username": USERNAME, "password": PASSWORD},
                            timeout=10)
            if r.status_code == 200:
                self.s.headers.update({"Authorization": f"Bearer {r.json()['access_token']}"})
                return True
            print(f"  ✕ Error de autenticación: {r.status_code}")
            return False
        except requests.ConnectionError:
            print("  ✕ No se puede conectar. ¿Está el servidor corriendo?")
            return False

    def get(self, ep):
        r = self.s.get(f"{BASE_URL}{ep}", timeout=15)
        return r.json() if r.status_code == 200 else []

    def post(self, ep, data):
        r = self.s.post(f"{BASE_URL}{ep}", json=data, timeout=15)
        if r.status_code in (200, 201): return r.json(), None
        try:    return None, r.json().get("detail", r.text)
        except: return None, r.text


# ── Leer Excel ─────────────────────────────────────────────
def leer_excel(filepath):
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]

    # Mapear columnas
    cols = {
        "factura":     find_col(headers, COL_FACTURA),
        "fecha":       find_col(headers, COL_FECHA),
        "apto":        find_col(headers, COL_APTO),
        "razon":       find_col(headers, COL_RAZON),
        "monto":       find_col(headers, COL_MONTO),
        "comentario":  find_col(headers, COL_COMENTARIO),
        "propietario": find_col(headers, COL_PROPIETARIO),
    }

    faltantes = [k for k in ["fecha", "apto", "monto"] if cols[k] is None]
    if faltantes:
        print(f"\n  ✕ Columnas obligatorias no encontradas: {faltantes}")
        print(f"  Columnas del archivo: {headers}")
        sys.exit(1)

    print(f"  Columnas detectadas:")
    for k, v in cols.items():
        if v is not None:
            print(f"    {k:15} → columna {v+1} ({headers[v]})")

    registros = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        def get(key):
            i = cols.get(key)
            return row[i] if i is not None and i < len(row) else None

        fecha = safe_date(get("fecha"))
        apto  = safe_str(get("apto"))
        monto = safe_float(get("monto"))

        if not fecha or not apto or not monto or monto <= 0:
            continue

        registros.append({
            "factura":     safe_str(get("factura")),
            "fecha":       fecha,
            "apto":        apto,
            "razon":       safe_str(get("razon")) or "Historial",
            "monto":       monto,
            "comentario":  safe_str(get("comentario")),
            "propietario": safe_str(get("propietario")),
        })

    return registros


# ── Validar y mostrar plan ─────────────────────────────────
def analizar(registros, unit_by_num, owner_by_unit):
    stats = {"ok": 0, "sin_unit": 0, "sin_owner": 0}
    aptos_faltantes = set()
    aptos_sin_owner = set()

    for r in registros:
        unit = unit_by_num.get(r["apto"])
        if not unit:
            stats["sin_unit"] += 1
            aptos_faltantes.add(r["apto"])
        elif not owner_by_unit.get(r["apto"]):
            stats["sin_owner"] += 1
            aptos_sin_owner.add(r["apto"])
        else:
            stats["ok"] += 1

    return stats, sorted(aptos_faltantes), sorted(aptos_sin_owner)


# ── Main ───────────────────────────────────────────────────
def main():
    print()
    print("═" * 62)
    print("  Importador de Historial de Pagos — Alianza Residencial")
    print("═" * 62)

    # Archivo a importar
    filepath = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_FILE
    if not os.path.exists(filepath):
        print(f"\n  ✕ Archivo no encontrado: {filepath}")
        print(f"  Uso: python importar_historial.py [archivo.xlsx]")
        sys.exit(1)

    print(f"\n  Archivo: {filepath}")
    print()

    # Leer Excel
    print("  Leyendo registros...")
    registros = leer_excel(filepath)
    print(f"  Registros válidos: {len(registros)}")

    if not registros:
        print("  No hay registros para importar.")
        sys.exit(0)

    # Estadísticas del archivo
    from collections import Counter
    años    = Counter(r["fecha"][:4] for r in registros)
    razones = Counter(r["razon"] for r in registros)
    aptos   = set(r["apto"] for r in registros)

    print(f"\n  Distribución por año:  {dict(sorted(años.items()))}")
    print(f"  Apartamentos únicos:   {len(aptos)}")
    print(f"  Razones/Conceptos:")
    for razon, cnt in razones.most_common():
        print(f"    {cnt:4d}x  {razon}")

    # Conectar
    print("\n  Conectando al servidor...")
    api = API()
    if not api.login():
        sys.exit(1)
    print("  ✓ Conectado")

    # Cargar catálogos
    units       = api.get("/units")
    unit_owners = api.get("/unit-owners/all")

    unit_by_num  = {u["unit_number"].strip(): u for u in units}
    owner_by_unit= {}
    for uo in unit_owners:
        if not uo.get("is_active"): continue
        for num, u in unit_by_num.items():
            if u["id"] == uo["unit_id"]:
                owner_by_unit[num] = uo["owner_id"]
                break

    print(f"  Unidades en sistema:   {len(units)}")

    # Analizar
    stats, faltantes, sin_owner = analizar(registros, unit_by_num, owner_by_unit)

    print()
    print("  ┌─────────────────────────────────────────────┐")
    print("  │  PLAN DE IMPORTACIÓN                        │")
    print("  ├─────────────────────────────────────────────┤")
    print(f"  │  Registros que se importarán:  {stats['ok']:>6}       │")
    print(f"  │  Apto no existe en sistema:    {stats['sin_unit']:>6}       │")
    print(f"  │  Apto sin propietario activo:  {stats['sin_owner']:>6}       │")
    print("  └─────────────────────────────────────────────┘")

    if faltantes:
        print(f"\n  ⚠  Aptos no registrados (se omitirán):")
        for a in faltantes[:15]:
            print(f"     · {a}")
        if len(faltantes) > 15:
            print(f"     ... y {len(faltantes)-15} más")

    if sin_owner:
        print(f"\n  ⚠  Aptos sin propietario activo (se omitirán):")
        for a in sin_owner[:10]:
            print(f"     · {a}")

    if stats["ok"] == 0:
        print("\n  No hay registros que se puedan importar.")
        sys.exit(0)

    print()
    resp = input(f"  ¿Importar {stats['ok']} registros? (s/n): ").strip().lower()
    if resp != "s":
        print("  Cancelado.")
        sys.exit(0)

    # ── Importar ───────────────────────────────────────────
    ok = 0; err = 0; omitidos = 0
    errores = []
    total = len(registros)
    t0 = time.time()

    print()
    for i, r in enumerate(registros, 1):
        if i % 50 == 0:
            print(f"  Progreso: {i}/{total} ({i*100//total}%)", end="\r")

        unit     = unit_by_num.get(r["apto"])
        owner_id = owner_by_unit.get(r["apto"])

        if not unit or not owner_id:
            omitidos += 1
            continue

        # Construir referencia combinando razón + comentario
        referencia = r["razon"]
        if r["comentario"]:
            referencia = f"{r['razon']} — {r['comentario']}"

        resultado, error = api.post("/finance/payment-complete", {
            "owner_id":       owner_id,
            "payment_date":   r["fecha"],
            "amount":         r["monto"],
            "invoice_number": r["factura"],
            "reference":      referencia,
            "applications":   [],   # historial puro — sin vincular a cargo
        })

        if resultado:
            ok += 1
        else:
            errores.append(f"Fac {r['factura']} | {r['apto']} | {r['fecha']} | ${r['monto']}: {error}")
            err += 1

    elapsed = time.time() - t0
    print()
    print()
    print("═" * 62)
    print("  RESULTADO")
    print("─" * 62)
    print(f"  ✓ Importados correctamente: {ok}")
    print(f"  ─ Omitidos (sin unidad):    {omitidos}")
    print(f"  ✕ Errores:                  {err}")
    print(f"  Tiempo: {elapsed:.1f}s")
    print("═" * 62)

    if errores:
        log_file = "historial_errores.txt"
        with open(log_file, "w", encoding="utf-8") as f:
            from datetime import datetime
            f.write(f"Errores de importación — {datetime.now()}\n\n")
            for e in errores: f.write(f"· {e}\n")
        print(f"\n  Log guardado en: {log_file}")
        print("  Primeros errores:")
        for e in errores[:8]: print(f"    · {e}")

    print()
    print("  Los pagos ya están disponibles en:")
    print("  · Finanzas → Historial de Cargos (por propietario)")
    print("  · Reportes → Por Propietario")
    print("  · Estado de Cuenta de cada propietario")


if __name__ == "__main__":
    main()
