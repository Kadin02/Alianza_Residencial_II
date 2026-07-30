"""
Script Utiles/importar_plantillas_simple.py

Importación en batch de los 3 archivos reales de Kadin, sin FIFO ni
PaymentApplication -- exactamente la misma lógica simple que ya usan
gas.html / imports.html (un Charge o un Payment suelto por fila), solo que
corrida de una vez en vez de fila por fila desde el navegador.

Este script es una HERRAMIENTA OPCIONAL. No se ejecuta solo, ni durante el
reset, ni al iniciar la app -- Kadin lo corre cuando decida usarlo, en vez
de (o además de) cargar los datos desde imports.html / gas.html.

Requiere que Owners/Units ya existan en la base de datos (reset_v2.py no
los toca). Si una unidad no se encuentra, la fila se omite y se reporta.

Fuentes (columnas reales, verificadas con openpyxl):

1. Migracion_Gas_Cargo.xlsx (hoja "Hoja1")
   MES, Apto, Propietario, Lectura Inicial, Lectura Final, Total a Pagar, YEAR, ID
   -> Charge(concept="Gas", description="Facturación de Gas - Lectura {LI} → {LF}",
             amount=round(Total a Pagar, 2), date_created=YEAR-MES-01, due_date=fin de mes)
   La columna ID es solo el orden de Kadin en Excel -- no se guarda.

2. Migracion_Gas_Pagos.xlsx (hoja "Hoja1")
   AÑO, MES, Fecha, Apartamento, Monto, Propietario, Concepto, # Factura, ID_Pago, ID_Cargo
   -> Payment(payment_date=Fecha, amount=round(Monto, 2), invoice_number=# Factura,
              concept=Concepto, reference=None)
   ID_Pago / ID_Cargo son solo bookkeeping de Kadin -- no se crea ningún
   PaymentApplication a partir de ID_Cargo, el pago no queda enlazado a un
   cargo específico en la BD.

3. Organizacion_Migracion 2024-2026.xlsx
   - Hoja "Cargos_Fijos_Atrasados" (año, concepto, monto_anual, propiedad):
     un Charge anual por unidad real (excluye "Por definir"/"A001") por cada
     fila -- misma lógica que ya tenía migrar_historial_2024_2025.py paso 4,
     sin FIFO.
   - Hoja "Pagos_Varios_Globales" (año, mes, fecha, Apartamento, factura,
     monto_pagado, concepto, Propiedad -- ojo: el nombre real de la hoja y
     el orden de columnas difieren de lo que se documentó originalmente;
     verificado con openpyxl contra el archivo real), FILTRANDO
     concepto != "Facturación de Gas" (esas filas son Gas, se manejan con
     los archivos 1 y 2 de arriba) -> Payment(payment_date=fecha,
     amount=round(monto_pagado, 2), invoice_number=factura, concept=concepto).
     "fecha" viene como número serial de Excel (sin formato de fecha en la
     celda), se convierte manualmente.
   "Clientes_Datos" y "Cargos_de_Gas" de este archivo YA NO se usan como
   fuente -- solo sirven de referencia manual si hace falta resolver un
   apartamento no encontrado.

Ningún Payment creado por este script genera OwnerCredit ni PaymentApplication
-- son filas sueltas, tal como ya se definió para gas.html/imports.html.

USO:
    python "Script Utiles/importar_plantillas_simple.py"              # dry-run
    python "Script Utiles/importar_plantillas_simple.py" --commit     # confirma
    python "Script Utiles/importar_plantillas_simple.py" --commit --skip-v2   # solo Gas
    python "Script Utiles/importar_plantillas_simple.py" --commit --skip-gas  # solo V2
"""

import argparse
import calendar
import os
import shutil
import sys
from datetime import date, datetime, timedelta
from decimal import Decimal

import openpyxl

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

_PROJECT_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
if _PROJECT_ROOT not in sys.path:
    sys.path.insert(0, _PROJECT_ROOT)

import app.main  # noqa: F401,E402  (registra todos los modelos)
from app.database import SessionLocal, engine  # noqa: E402
from app.models.unit import Unit  # noqa: E402
from app.models.unit_owner import UnitOwner  # noqa: E402
from app.models.charge import Charge  # noqa: E402
from app.models.payment import Payment  # noqa: E402

GAS_CARGO_PATH = os.path.join(os.path.dirname(__file__), "Migracion_Gas_Cargo.xlsx")
GAS_PAGOS_PATH = os.path.join(os.path.dirname(__file__), "Migracion_Gas_Pagos.xlsx")
V2_PATH        = os.path.join(os.path.dirname(__file__), "Organizacion_Migracion 2024-2026.xlsx")

CENT = Decimal("0.01")
NON_REAL_APTOS = {"Por definir", "A001"}


def to_dec(value) -> Decimal:
    return Decimal(str(value)).quantize(CENT)


def excel_serial_to_date(value):
    """'fecha' en Pagos_Varios_Globales viene como número serial de Excel
    (celda sin formato de fecha), no como datetime -- se convierte a mano
    con el epoch estándar de Excel (1899-12-30)."""
    if value is None:
        return None
    if hasattr(value, "date"):
        return value.date()
    if isinstance(value, (int, float)):
        return date(1899, 12, 30) + timedelta(days=int(value))
    return value


def read_sheet_rows(wb, sheet_name):
    """Descarta filas completamente vacías (Excel suele traer un rango
    usado más grande que los datos reales)."""
    ws = wb[sheet_name]
    rows = ws.iter_rows(min_row=2, values_only=True)
    return [r for r in rows if any(v is not None for v in r)]


def backup_database():
    if engine.dialect.name != "sqlite":
        print(f"[ERROR] DATABASE_URL apunta a '{engine.dialect.name}', no a SQLite.")
        print("Respalda manualmente (ej. pg_dump) antes de correr --commit contra Postgres.")
        sys.exit(1)
    db_path = engine.url.database
    if not db_path or not os.path.exists(db_path):
        print(f"[ERROR] No se encontró el archivo de base de datos en '{db_path}'")
        sys.exit(1)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = f"{db_path}.bak_{timestamp}_pre_importar_plantillas"
    shutil.copy2(db_path, backup_path)
    print(f"[OK] Backup creado: {backup_path}")


class Report:
    def __init__(self):
        self.gas_charges = 0
        self.gas_payments = 0
        self.v2_charges = 0
        self.v2_payments = 0
        self.omitted = []

    def omit(self, source, row, reason):
        self.omitted.append({"source": source, "row": row, "reason": reason})

    def print_summary(self):
        print("\n" + "=" * 70)
        print("REPORTE — Importación de plantillas")
        print("=" * 70)
        print(f"Cargos de Gas creados:   {self.gas_charges}")
        print(f"Pagos de Gas creados:    {self.gas_payments}")
        print(f"Cargos V2 creados:       {self.v2_charges}")
        print(f"Pagos V2 creados:        {self.v2_payments}")
        print(f"\nFilas omitidas: {len(self.omitted)}")
        for o in self.omitted[:30]:
            print(f"  [{o['source']}] {o['row']} -> {o['reason']}")
        if len(self.omitted) > 30:
            print(f"  ... y {len(self.omitted) - 30} más")
        print("=" * 70)


def import_gas_cargos(db, report, unit_by_apto):
    if not os.path.exists(GAS_CARGO_PATH):
        print(f"[SKIP] No se encontró {GAS_CARGO_PATH}")
        return
    wb = openpyxl.load_workbook(GAS_CARGO_PATH, data_only=True)
    for row in read_sheet_rows(wb, "Hoja1"):
        mes, apto, propietario, li, lf, total, year, _id = row
        if apto is None or mes is None or year is None or total is None:
            report.omit("Migracion_Gas_Cargo", row, "campos obligatorios vacíos")
            continue

        unit = unit_by_apto.get(str(apto))
        if unit is None:
            report.omit("Migracion_Gas_Cargo", row, f"apartamento '{apto}' no encontrado")
            continue

        mes, year = int(mes), int(year)
        due_day = calendar.monthrange(year, mes)[1]
        charge = Charge(
            unit=unit,
            concept="Gas",
            description=f"Facturación de Gas - Lectura {li} → {lf}",
            amount=to_dec(total),
            status="PENDIENTE",
            date_created=date(year, mes, 1),
            due_date=date(year, mes, due_day),
        )
        db.add(charge)
        report.gas_charges += 1


def import_gas_pagos(db, report, unit_by_apto, active_owner_by_unit):
    if not os.path.exists(GAS_PAGOS_PATH):
        print(f"[SKIP] No se encontró {GAS_PAGOS_PATH}")
        return
    wb = openpyxl.load_workbook(GAS_PAGOS_PATH, data_only=True)
    for row in read_sheet_rows(wb, "Hoja1"):
        anio, mes, fecha, apto, monto, propietario, concepto, factura, id_pago, id_cargo = row
        if apto is None or fecha is None or monto is None:
            report.omit("Migracion_Gas_Pagos", row, "campos obligatorios vacíos")
            continue

        unit = unit_by_apto.get(str(apto))
        if unit is None:
            report.omit("Migracion_Gas_Pagos", row, f"apartamento '{apto}' no encontrado")
            continue

        owner = active_owner_by_unit.get(unit.id)
        if owner is None:
            report.omit("Migracion_Gas_Pagos", row, f"unidad '{apto}' sin propietario activo")
            continue

        fecha_pago = fecha.date() if hasattr(fecha, "date") else fecha
        payment = Payment(
            property=unit.property,
            owner=owner,
            payment_date=fecha_pago,
            amount=to_dec(monto),
            total_amount=to_dec(monto),
            invoice_number=str(factura) if factura is not None else None,
            reference=None,
            concept=str(concepto) if concepto else "Facturación de Gas",
        )
        db.add(payment)
        report.gas_payments += 1


def import_v2_cargos_fijos(db, report, unit_by_apto):
    """Mismo patrón que migrar_historial_2024_2025.py paso 4, sin FIFO:
    un Charge anual por unidad real (excluye NON_REAL_APTOS) por cada fila
    de Cargos_Fijos_Atrasados."""
    wb = openpyxl.load_workbook(V2_PATH, data_only=True)
    if "Cargos_Fijos_Atrasados" not in wb.sheetnames:
        print("[SKIP] Hoja 'Cargos_Fijos_Atrasados' no encontrada")
        return

    annual_defs = []
    for row in read_sheet_rows(wb, "Cargos_Fijos_Atrasados"):
        anio, concepto, monto_anual, _propiedad = row
        if anio is None or concepto is None or monto_anual is None:
            continue
        anio = int(anio)
        if concepto.strip() == "Mantenimiento":
            match_concepto, display_desc = "Mantenimiento", "Mantenimiento"
        elif "Ascensor" in concepto:
            match_concepto, display_desc = "Ascensor", "Cuota Ascensor"
        else:
            report.omit("Cargos_Fijos_Atrasados", row, f"concepto '{concepto}' no reconocido")
            continue
        annual_defs.append((match_concepto, display_desc, anio, to_dec(monto_anual)))

    for apto, unit in unit_by_apto.items():
        if apto in NON_REAL_APTOS:
            continue
        for match_concepto, display_desc, anio, monto in annual_defs:
            charge = Charge(
                unit=unit,
                concept=match_concepto,
                description=f"{display_desc} {anio}",
                amount=monto,
                status="PENDIENTE",
                date_created=date(anio, 1, 1),
                due_date=date(anio, 12, 31),
            )
            db.add(charge)
            report.v2_charges += 1


def import_v2_pagos_global(db, report, unit_by_apto, active_owner_by_unit):
    """Hoja real 'Pagos_Varios_Globales' (año, mes, fecha, Apartamento,
    factura, monto_pagado, concepto, Propiedad), filtrando fuera
    'Facturación de Gas' (esas filas se importan aparte con
    Migracion_Gas_Pagos.xlsx)."""
    wb = openpyxl.load_workbook(V2_PATH, data_only=True)
    if "Pagos_Varios_Globales" not in wb.sheetnames:
        print("[SKIP] Hoja 'Pagos_Varios_Globales' no encontrada")
        return

    for row in read_sheet_rows(wb, "Pagos_Varios_Globales"):
        anio, mes, fecha, apto, factura, monto_pagado, concepto, _propiedad = row

        if concepto == "Facturación de Gas":
            continue  # Gas se importa aparte, con sus propias plantillas

        if apto is None or fecha is None or monto_pagado is None:
            report.omit("Pagos_Varios_Globales", row, "campos obligatorios vacíos")
            continue

        unit = unit_by_apto.get(str(apto))
        if unit is None:
            report.omit("Pagos_Varios_Globales", row, f"apartamento '{apto}' no encontrado")
            continue

        owner = active_owner_by_unit.get(unit.id)
        if owner is None:
            report.omit("Pagos_Varios_Globales", row, f"unidad '{apto}' sin propietario activo")
            continue

        fecha_pago = excel_serial_to_date(fecha)
        payment = Payment(
            property=unit.property,
            owner=owner,
            payment_date=fecha_pago,
            amount=to_dec(monto_pagado),
            total_amount=to_dec(monto_pagado),
            invoice_number=str(factura) if factura is not None else None,
            reference=None,
            concept=str(concepto) if concepto else None,
        )
        db.add(payment)
        report.v2_payments += 1


def run_import(commit: bool, skip_gas: bool, skip_v2: bool):
    backup_database()

    report = Report()
    db = SessionLocal()
    try:
        units = db.query(Unit).all()
        unit_by_apto = {u.unit_number: u for u in units}

        active_relations = db.query(UnitOwner).filter(UnitOwner.is_active == True).all()
        active_owner_by_unit = {r.unit_id: r.owner for r in active_relations}

        if not skip_gas:
            import_gas_cargos(db, report, unit_by_apto)
            import_gas_pagos(db, report, unit_by_apto, active_owner_by_unit)

        if not skip_v2:
            if os.path.exists(V2_PATH):
                import_v2_cargos_fijos(db, report, unit_by_apto)
                import_v2_pagos_global(db, report, unit_by_apto, active_owner_by_unit)
            else:
                print(f"[SKIP] No se encontró {V2_PATH}")

        db.flush()
        report.print_summary()

        if commit:
            db.commit()
            print("\n[OK] COMMIT realizado - los datos quedaron guardados en la base de datos real.")
        else:
            db.rollback()
            print("\n[DRY-RUN] Se hizo ROLLBACK, no se guardó nada. Corre con --commit para confirmar.")

    except Exception:
        db.rollback()
        print("\n[ERROR] Se hizo ROLLBACK completo, no quedó nada a medias.")
        raise
    finally:
        db.close()


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Importación en batch de las plantillas reales (Gas + V2), sin FIFO")
    parser.add_argument("--commit", action="store_true", help="Confirma los cambios (por defecto es dry-run)")
    parser.add_argument("--skip-gas", action="store_true", help="No importar Migracion_Gas_Cargo.xlsx / Migracion_Gas_Pagos.xlsx")
    parser.add_argument("--skip-v2", action="store_true", help="No importar Organizacion_Migracion 2024-2026.xlsx")
    args = parser.parse_args()

    run_import(commit=args.commit, skip_gas=args.skip_gas, skip_v2=args.skip_v2)
